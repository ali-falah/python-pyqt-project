from PyQt4.QtGui import *
from PyQt4.QtCore import *
from PyQt4 import QtGui
from PyQt4 import QtCore
import  sys
from os import path
import time
import openpyxl
from openpyxl import Workbook
from PyQt4.uic import loadUiType
import threading
import sqlite3
import webbrowser


Form_class,_= loadUiType(path.join(path.dirname(__file__),"hotel.ui"))


class main(QMainWindow,Form_class):
    db = sqlite3.connect("hotel_info.db")
    cur=db.cursor()


    def __init__(self):
        QMainWindow.__init__(self)
        self.setupUi(self)
        self.stackedWidget.setCurrentIndex(0)
        self.setFixedSize(802,612)
        self.dateTimeEdit.setDateTime(QtCore.QDateTime.currentDateTime())
        self.dateTimeEdit_2.setDateTime(QtCore.QDateTime.currentDateTime())
        self.dateTimeEdit_3.setDateTime(QtCore.QDateTime.currentDateTime())
        self.handel_buttons()

    def handel_buttons(self):
        self.pushButton_5.clicked.connect(self.login)
        self.exitbut.clicked.connect(self.exit)
        self.pushButton_4.clicked.connect(self.logout)
        self.submitbut.clicked.connect(self.submituser)
        self.pushButton_2.clicked.connect(self.Delete_db)
        self.pushButton_3.clicked.connect(self.Update_into_db)
        self.pushButton.clicked.connect(self.show_list)
        self.pushButton_6.clicked.connect(self.Manage_acounts_window)
        self.exitbut_2.clicked.connect(self.exit)
        self.exitbut_3.clicked.connect(self.return_window)
        self.pushButton_9.clicked.connect(self.ADD_Employee)
        self.pushButton_11.clicked.connect(self.show_list_2)
        self.pushButton_8.clicked.connect(self.Remove_Employee)
        self.pushButton_10.clicked.connect(self.Update_Employee)
        self.pushButton_7.clicked.connect(self.search_costumer)
        self.pushButton_12.clicked.connect(self.search_Employee)
        self.pushButton_15.clicked.connect(self.Clear_2)
        self.pushButton_16.clicked.connect(self.Clear)
        self.pushButton_14.clicked.connect(self.Export)
        self.pushButton_13.clicked.connect(self.Export_2)
        self.pushButton_17.clicked.connect(self.About_Window)

    def About_Window(self):
        win=About_Window()
        win.setFixedSize(703,648)
        win.exec_()


    def Export_2(self):
        Sheet_name = r"employee_info.xlsx"
        rb = openpyxl.load_workbook(Sheet_name)
        sheet = rb.active

        data = self.db.execute("SELECT * FROM Users")
        i=2
        for r in data:
            sheet.cell(row=i, column=1 , value = r[0] )
            sheet.cell(row=i, column=2 , value = r[1] )
            sheet.cell(row=i, column=3 , value = r[2] )
            sheet.cell(row=i, column=4 , value = r[3] )
            sheet.cell(row=i, column=5 , value = r[4] )
            sheet.cell(row=i, column=6 , value = r[5] )
            i+=1


        rb.save(Sheet_name)
        self.statusbar.showMessage('DATA EXPORTED SUCCESSFULLY :)')
        self.sleeper()


    def Export(self):
        Sheet_name = r"costumer_info.xlsx"
        rb = openpyxl.load_workbook(Sheet_name)
        sheet = rb.active


        sheet.cell(row=1, column=1 , value = "ID")
        sheet.cell(row=1, column=2 , value ='NAME' )
        sheet.cell(row=1, column=3 , value ='STATUS' )
        sheet.cell(row=1, column=4 , value ='COMMENT' )
        sheet.cell(row=1, column=5 , value ='SERVICE_TYPE' )
        sheet.cell(row=1, column=6 , value ='STAY_DAYS' )
        sheet.cell(row=1, column=7 , value ='COST' )
        sheet.cell(row=1, column=8 , value ='PAID' )

        data = self.db.execute("SELECT * FROM costumer")
        i=2
        for r in data:
            sheet.cell(row=i, column=1 , value = r[0] )
            sheet.cell(row=i, column=2 , value = r[1] )
            sheet.cell(row=i, column=3 , value = r[2] )
            sheet.cell(row=i, column=4 , value = r[3] )
            sheet.cell(row=i, column=5 , value = r[4] )
            sheet.cell(row=i, column=6 , value = r[5] )
            sheet.cell(row=i, column=7 , value = r[6] )
            sheet.cell(row=i, column=8 , value = r[7] )
            i+=1
        rb.save(Sheet_name)
        self.statusbar.showMessage('DATA EXPORTED SUCCESSFULLY :)')
        self.sleeper()

    def Clear(self):
        self.LINEEDIT_up_del.clear()
        self.nameline.clear()
        self.familyline.clear()
        self.tedit.clear()
        self.comboBox.setCurrentIndex(0)
        self.spinBox.setValue(0)
        self.rdbut1.setChecked(False)
        self.rdbut2.setChecked(False)
        self.cost_line_edit.clear()


    def search_Employee(self):
        try:
            self.tableWidget_2.setRowCount(0)
            name = self.lineEdit_7.text()
            if name.isalpha()==True :
                c = self.db.execute("SELECT * FROM Users WHERE username = ? ",(name,))
                self.lineEdit_7.clear()

            else:
                id = self.lineEdit_5.text()
                c  = self.db.execute("SELECT * FROM Users WHERE ID = ? ",(int(id),))
                self.lineEdit_5.clear()

            i=0
            for r in c:
                self.tableWidget_2.insertRow(i)
                self.tableWidget_2.setItem(i,0,QTableWidgetItem(str(r[0])))
                self.lineEdit_5.setText(str(r[0]))
                self.tableWidget_2.setItem(i,1,QTableWidgetItem(r[1]))
                self.lineEdit_3.setText(r[1])
                self.tableWidget_2.setItem(i,2,QTableWidgetItem(str(r[2])))
                self.lineEdit_4.setText(str(r[2]))
                self.lineEdit_6.setText(str(r[2]))
                self.tableWidget_2.setItem(i,3,QTableWidgetItem(r[3]))
                if r[3]=='Admin':
                    self.comboBox_2.setCurrentIndex(0)
                elif r[3]=='Recipent':
                    self.comboBox_2.setCurrentIndex(1)
                elif r[3]=='Room Servant':
                    self.comboBox_2.setCurrentIndex(2)
                else:
                    self.comboBox_2.setCurrentIndex(3)

                self.tableWidget_2.setItem(i,4,QTableWidgetItem(str(r[4])))
                self.lineEdit_8.setText(str(r[4]))
                self.tableWidget_2.setItem(i,5,QTableWidgetItem(r[5]))
                if r[5]=='yes':
                    self.radioButton_4.setChecked(True)
                else:
                    self.radioButton_3.setChecked(True)
                i+=1

        except Exception:
            self.statusbar.showMessage('DIDNOT FIND IT ')
            self.sleeper()


    def search_costumer(self):
        self.tableWidget.setRowCount(0)
        name=self.nameline.text() + " " + self.familyline.text()
        try:
            if name!="" and self.LINEEDIT_up_del.text()=="":
                c = self.db.execute("SELECT * FROM costumer WHERE Name = ? ",(name,))
            else:
                c = self.db.execute("SELECT * FROM costumer WHERE Id = ? ",(self.LINEEDIT_up_del.text()))
            i=0
            for r in c:
                self.tableWidget.insertRow(i)
                self.tableWidget.setItem(i,0,QTableWidgetItem(str(r[0])))
                self.LINEEDIT_up_del.setText(str(r[0]))
                self.tableWidget.setItem(i,1,QTableWidgetItem(r[1]))
                fn , ln = r[1].split(" ")
                self.nameline.setText(fn)
                self.familyline.setText(ln)
                self.tableWidget.setItem(i,2,QTableWidgetItem(r[2]))
                if r[2]=="Married":
                    self.rdbut2.setChecked(True)
                else:
                    self.rdbut1.setChecked(True)
                self.tableWidget.setItem(i,3,QTableWidgetItem(r[3]))
                self.tedit.insertPlainText(r[3])
                self.tableWidget.setItem(i,4,QTableWidgetItem(r[4]))
                if r[4]=='Ordinary room\s':
                    self.comboBox.setCurrentIndex(0)
                elif r[4]=='Sweet':
                    self.comboBox.setCurrentIndex(1)
                else:
                    self.comboBox.setCurrentIndex(2)
                self.tableWidget.setItem(i,5,QTableWidgetItem(str(r[5])))
                self.spinBox.setValue(r[5])
                self.tableWidget.setItem(i,6,QTableWidgetItem(str(r[6])))
                self.cost_line_edit.setText(str(r[6]))
                self.tableWidget.setItem(i,7,QTableWidgetItem(r[7]))
                if r[7]=='yes':
                    self.radioButton.setChecked(True)
                else:
                    self.radioButton_2.setChecked(True)
                i+=1
        except Exception:
            self.statusbar.showMessage('CANNOT SEARCH CHECK ID AND NAME FIELDS ')
            self.sleeper()

    def Clear_2(self):
        self.comboBox_2.setCurrentIndex(0)
        self.radioButton_3.setChecked(False)
        self.radioButton_4.setChecked(False)
        self.lineEdit_3.clear()
        self.lineEdit_4.clear()
        self.lineEdit_6.clear()
        self.lineEdit_8.clear()
        self.lineEdit_5.clear()


    def Update_Employee(self):
        try:
            if self.lineEdit_3.text()!="":
                if self.lineEdit_4.text()!="":
                    if self.lineEdit_4.text()==self.lineEdit_6.text():
                        if self.lineEdit_8.text()!="":
                            user=self.lineEdit_3.text()
                            password=self.lineEdit_4.text()
                            salary=self.lineEdit_8.text()
                            if self.comboBox_2.currentIndex()==0:
                                type='Admin'
                            elif self.comboBox_2.currentIndex()==1:
                                type='Recipent'
                            elif self.comboBox_2.currentIndex()==2:
                                type='Room servant'
                            else:
                                type='Cleaner'

                            if self.radioButton_4.isChecked()==True:
                                paid='yes'
                            else:
                                paid='no'

                            id = self.lineEdit_5.text()


                            sql = '''UPDATE Users
                                    SET username = ? ,
                                    password = ? ,
                                    Employee_type = ? ,
                                    SALARY = ? ,
                                    paidORnot = ?
                                    Where ID =? '''
                            self.cur.execute(sql , (user , int(password) , type ,int(salary), paid , int(id) ))
                            self.db.commit()
                            self.statusbar.showMessage('DATABASE UPDATED SUCCESSFULLY :)')
                            self.sleeper()
                            self.Clear_2()

                        else:
                            self.statusbar.showMessage('SALARY CANNOT BE EMPETY')
                            self.sleeper()

                    else:
                        self.statusbar.showMessage('PASSWORDS ISNOT THE SAME RETYPE IT AGAIN')
                        self.sleeper()
                else:
                    self.statusbar.showMessage('PASSWORD FIELED CANNOT BE EMPETY')
                    self.sleeper()
            else:
                self.statusbar.showMessage('USERNAME FIELED CANNOT BE EMPETY')
        except Exception:
            self.statusbar.showMessage('COULDNOT UPDATE CHECK IF PASSWORD AND USERNAME IS NOT USED ? ')
            self.sleeper
    def Remove_Employee(self):
        if self.lineEdit_5.text() !="":
            try:
                msg = QMessageBox()
                msg.setIcon( QtGui.QMessageBox.Critical )
                msg.setText( "YOU ARE ABOUT TO REMOVE AN EMPLOYEE" )
                msg.setWindowTitle("Hotel MANAGEMENT")
                msg.setWindowIcon( QIcon("DELETE.png"))
                msg.setInformativeText( "ARE YOU SURE ? " )
                msg.addButton( QtGui.QMessageBox.Yes )
                msg.addButton( QtGui.QMessageBox.No )

                msg.setDefaultButton( QtGui.QMessageBox.No )
                ret = msg.exec_()

                if ret == QtGui.QMessageBox.Yes:
                    user = self.lineEdit_5.text()
                    self.db.execute("DELETE FROM Users WHERE ID = ? ",(user))
                    self.db.commit()
                    self.statusbar.showMessage('USER HAS DELETED SSUCCESSFULLY :)')
                    self.sleeper()
                    return
                else:
                    print( "No" )
                    return

            except Exception:
                self.statusbar.showMessage('USER HASNOT DELETED CHECK THE ID NUMBER AND TRY AGAIN :(')
                self.sleeper()
        else:
            self.statusbar.showMessage('SORRY THE ID NUMBER FIELED CANNOT BE EMPETY ')
            self.sleeper()

    def show_list_2(self):
        self.tableWidget_2.setRowCount(0)

        data = self.db.execute('SELECT * FROM Users')
        i=0
        for r in data:
            self.tableWidget_2.insertRow(i)
            self.tableWidget_2.setItem(i,0,QTableWidgetItem(str(r[0])))
            self.tableWidget_2.setItem(i,1,QTableWidgetItem(r[1]))
            self.tableWidget_2.setItem(i,2,QTableWidgetItem(str(r[2])))
            self.tableWidget_2.setItem(i,3,QTableWidgetItem(r[3]))
            self.tableWidget_2.setItem(i,4,QTableWidgetItem(str(r[4])))
            self.tableWidget_2.setItem(i,5,QTableWidgetItem(r[5]))
            i+=1


    def ADD_Employee(self):
        try:
            if self.lineEdit_3.text()!="" or self.lineEdit_4.text()!="":

                if self.lineEdit_6.text()!="":
                    if self.lineEdit_4.text()==self.lineEdit_6.text():
                        password=self.lineEdit_4.text()
                        self.lineEdit_4.clear()
                        self.lineEdit_6.clear()

                        int(password)
                        user=self.lineEdit_3.text()
                        self.lineEdit_3.clear()


                        if self.comboBox_2.currentIndex()==0:
                            employee_type='Admin'
                        elif self.comboBox_2.currentIndex()==1:
                            employee_type='Recipent'
                        elif self.comboBox_2.currentIndex()==2:
                            employee_type='Room Servant'
                        else:
                            employee_type='Cleaner'
                        self.comboBox_2.setCurrentIndex(0)

                        paid='NO'
                        if self.radioButton_4.isChecked()==True:
                            paid='yes'
                            self.radioButton_4.setChecked(False)

                        if self.lineEdit_8.text()!="":
                            salary=int(self.lineEdit_8.text())
                            self.lineEdit_8.clear()
                            self.db.execute("INSERT INTO Users(username,password,EMPLOYEE_TYPE,SALARY,paidORnot) VALUES(?,?,?,?,?)",(user,int(password),employee_type,salary,paid))
                        else:
                            self.statusbar.showMessage('PLEASE ENTER EMPLOYEE SALARY')
                            self.sleeper()
                        self.db.commit()
                        self.statusbar.showMessage('USER HAS ADDED SUCCESSFULLY :)')
                        self.sleeper()
                    else:
                        self.statusbar.showMessage('PASSWORD ISNOT THE SAME TRY AGAIN AND CHECK THE CASPLOCK KEY ?')
                        self.sleeper()

                else:
                    self.statusbar.showMessage('PLEASE CONFIRM THE PASSSWORD ')
                    self.sleeper()
            else:
                self.statusbar.showMessage('USERNAME OR PASSWORD FIELD CANNOT BE EMPETY')
                self.sleeper()
        except Exception:
            self.statusbar.showMessage('COLUDNOT ADD EMLPOYEE CHECK USENAME AND PASSSWORD ISNOT USED BY OTHER ENPLOYEE')
            self.sleeper()

    def return_window(self):
        self.stackedWidget.setCurrentIndex(1)
        self.stackedWidget.resize(1490,722)
        self.setFixedSize(1490,780)
        self.move(100,100)
        self.statusbar.showMessage('DEVELOPED BY ALI FALAH ')


    def Manage_acounts_window(self):
        self.stackedWidget.setCurrentIndex(2)
        self.stackedWidget.resize(1401,751)
        self.setFixedSize(1406,795)
        self.move(100,100)
        self.statusbar.showMessage('DEVELOPED BY ALI FALAH ')


    def login(self):
        userData = self.db.execute('SELECT * FROM Users')
        userline = self.lineEdit.text()
        passline = self.lineEdit_2.text()
        usertest  = False
        passtest  = False
        admintest = True
        for r in userData:
            if userline == r[1]:
                usertest = True
                if passline == str(r[2]):
                    passtest == True
                    self.stackedWidget.setCurrentIndex(1)
                    self.stackedWidget.resize(1490,722)
                    self.setFixedSize(1490,780)
                    self.move(100,100)
                    self.statusbar.showMessage('DEVELOPED BY ALI FALAH ____WELCOM ADMIN____ ')
                    if r[3]!='Admin':
                        admintest == False
                        self.pushButton_6.setEnabled(False)
                        self.statusbar.showMessage('DEVELOPED BY ALI FALAH ____WELCOMe SIR____ ')

                else:
                    self.statusbar.showMessage('PASSWORD INCORRECT RETYPE IT AGAIN')
                    self.lineEdit_2.clear()
                    self.sleeper()
        if usertest == False :
                self.statusbar.showMessage('USERNAME IS NOT FOUND ')
                self.sleeper()
        if admintest == False :
            self.sleeper()
            self.statusbar.showMessage('YOU HAVE NO PRIVILGES TO MANAGE ACCOUNTS ')


    def logout(self):
        msg = QMessageBox()
        msg.setIcon( QtGui.QMessageBox.Critical )
        msg.setText( "YOU ARE ABOUT TO LOGOUT " )
        msg.setWindowTitle("Hotel Reservaion")
        msg.setWindowIcon( QIcon("DELETE.png"))
        msg.setInformativeText( "ARE YOU SURE ?" )
        msg.addButton( QtGui.QMessageBox.Yes )
        msg.addButton( QtGui.QMessageBox.No )

        msg.setDefaultButton( QtGui.QMessageBox.No )
        ret = msg.exec_()

        if ret == QtGui.QMessageBox.Yes:
            self.stackedWidget.setCurrentIndex(0)
            self.stackedWidget.resize(802,571)
            self.setFixedSize(802,612)
            self.move(100,100)
            self.statusbar.showMessage('DEVELOPED BY ALI FALAH ')
            return
        else:
            print( "No" )
            return



    def exit(self):
        msg = QMessageBox()
        msg.setIcon( QtGui.QMessageBox.Critical )
        msg.setText( "YOU ARE ABOUT TO CLOSE THE APP" )
        msg.setWindowTitle("Hotel Reservaion")
        msg.setWindowIcon( QIcon("DELETE.png"))
        msg.setInformativeText( "ARE YOU SURE ? :(" )
        msg.addButton( QtGui.QMessageBox.Yes )
        msg.addButton( QtGui.QMessageBox.No )

        msg.setDefaultButton( QtGui.QMessageBox.No )
        ret = msg.exec_()

        if ret == QtGui.QMessageBox.Yes:
            sys.exit()
            return
        else:
            print( "No" )
            return

    def ADD_into_db(self,name,gender,comment,service,days,cost,paidornot):
        self.db.row_factory=sqlite3.Row
        int(days)
        int(cost)
        try:
            self.db.execute("INSERT INTO costumer( Name , gender , comment , service_type , stay_days , COST, paidORnot ) VALUES(?,?,?,?,?,?,?)",(name,gender,comment,service,days,cost,paidornot))
            self.db.commit()
            self.statusbar.showMessage('Add User Task Finished Successfully :)')
            self.sleeper()
        except Exception:
            self.statusbar.showMessage('COLUDNOT FINISH THE TASK COZ NAME IS EXISTED PLEASE CHOOSE ANOTHER :)')
            self.sleeper()



    def submituser(self):
        if self.nameline.text()!="" and self.familyline.text()!="":
            if self.rdbut1.isChecked()==True or self.rdbut2.isChecked()==True:
                name=self.nameline.text() + " " + self.familyline.text()
                self.nameline.clear()
                self.familyline.clear()
                if self.rdbut1.isChecked() == True:
                    status="Single"
                else:
                    status="Married"
                comment=self.tedit.toPlainText()
                self.tedit.clear()
                service='test'

                if self.comboBox.currentIndex()==0:
                    service='Ordinary room\s'
                elif self.comboBox.currentIndex()==1:
                    service='Sweet'
                else:
                    service='whole wing'
                self.comboBox.setCurrentIndex(0)
                days = self.spinBox.value()

                cost=self.cost_line_edit.text()
                paidornot='No'
                if self.radioButton.isChecked()== True:
                    paidornot='yes'
                if self.cost_line_edit.text() !="":
                    self.ADD_into_db(name,status,comment,service,days,cost,paidornot)
                    self.cost_line_edit.clear()
                else:
                    self.statusbar.showMessage('COST ARE REQUIRED PLEASE ENTER IT :)')
                    self.sleeper()
            else:
                self.statusbar.showMessage('PLEASE CHOOSE THE SOCAIL STATUS')
                self.sleeper()
        else:
            self.statusbar.showMessage('PLEASE FULL THE INFO ABOVE')
            self.sleeper()


    def show_list(self):

        self.tableWidget.setRowCount(0)

        data = self.db.execute('SELECT * FROM costumer')
        i=0
        for r in data:
            self.tableWidget.insertRow(i)
            self.tableWidget.setItem(i,0,QTableWidgetItem(str(r[0])))
            self.tableWidget.setItem(i,1,QTableWidgetItem(r[1]))
            self.tableWidget.setItem(i,2,QTableWidgetItem(r[2]))
            self.tableWidget.setItem(i,3,QTableWidgetItem(r[3]))
            self.tableWidget.setItem(i,4,QTableWidgetItem(r[4]))
            self.tableWidget.setItem(i,5,QTableWidgetItem(str(r[5])))
            self.tableWidget.setItem(i,6,QTableWidgetItem(str(r[6])))
            self.tableWidget.setItem(i,7,QTableWidgetItem(r[7]))
            i+=1

    def Delete_db(self):
        if self.LINEEDIT_up_del.text()!="":
            try:
                msg = QMessageBox()
                msg.setIcon( QtGui.QMessageBox.Critical )
                msg.setText( "YOU ARE ABOUT TO DELETE A RESERVTION" )
                msg.setWindowTitle("Hotel Reservaion")
                msg.setWindowIcon( QIcon("DELETE.png"))
                msg.setInformativeText( "ARE YOU SURE ?" )
                msg.addButton( QtGui.QMessageBox.Yes )
                msg.addButton( QtGui.QMessageBox.No )

                msg.setDefaultButton( QtGui.QMessageBox.No )
                ret = msg.exec_()

                if ret == QtGui.QMessageBox.Yes:
                    id = self.LINEEDIT_up_del.text()
                    int(id)
                    self.db.execute("Delete From costumer Where ID = ? ",(id))
                    self.nameline.clear()
                    self.familyline.clear()
                    self.db.commit()
                    self.statusbar.showMessage('Record Has Deleted Successfully :)')
                    return
                else:
                    print( "No" )
                    return

            except Exception:
                self.statusbar.showMessage('Record Hasnot Deleted Successfully :( Check The ID Number')
                self.sleeper()
        else:
            self.statusbar.showMessage('ID NUMBER FIELED CAN NOT BE EMPETY')
            self.sleeper()


    def Update_into_db(self):
        try:
            name=self.nameline.text() + " " + self.familyline.text()
            status='test'
            if self.rdbut1.isChecked() == True:
                status="Single"
            if self.rdbut2.isChecked()==True:
                status="Married"
            comment=self.tedit.toPlainText()

            service='test'
            if self.comboBox.currentIndex()==0:
                service='Ordiary room\s'
            elif self.comboBox.currentIndex()==1:
                service='Sweet'
            else:
                service='whole wing'

            days=int(self.spinBox.value())
            cost=self.cost_line_edit.text()
            if self.radioButton.isChecked()== True:
                paidornot='yes'
            else:
                paidornot='No'

            id=int(self.LINEEDIT_up_del.text())

            if name!="":
                if status!='test':
                    if cost!="":
                        if id!="":
                            sql='''UPDATE COSTUMER
                            SET Name= ? ,
                            gender= ? ,
                            comment = ? ,
                            service_type = ?,
                            stay_days = ? ,
                            COST = ? ,
                            paidORnot = ?
                            where ID = ? '''
                            self.cur.execute( sql ,(name,status, comment, service,days,int(cost),paidornot,id))
                            self.db.commit()
                            self.statusbar.showMessage('Database Updated Successfully :)')
                            self.Clear()
                            self.sleeper()
                        else:
                            self.statusbar.showMessage('ID NUMBER FIELD CANNOT BE EMPETY')
                    else:
                        self.statusbar.showMessage('ENTER COST PLEASE')
                        self.sleeper()
                else:
                    self.statusbar.showMessage('PLEASE CHOOSE GENDER')
                    self.sleeper()
            else:
                self.statusbar.showMessage('NAME OR FAMILY NAME CANNOT BE EMPETY')
                self.sleeper()


        except Exception:
            self.statusbar.showMessage('Database ColudNot Updated CHECK ID OR NAME ISNOT EXISTED :(')
            self.sleeper()



######################################################
    def sleep_function(self):
        time.sleep(7)
        self.statusbar.showMessage('')
    def sleeper(self):
        t=threading.Thread(target=self.sleep_function)
        t.start()
#######################################################
Window_class,_= loadUiType(path.join(path.dirname(__file__),"about_window.ui"))

class About_Window(QDialog,Window_class):

    def __init__(self):
        QDialog.__init__(self)
        self.setupUi(self)
        self.handel_buttons()

    def handel_buttons(self):
        self.pushButton_2.clicked.connect(self.facebook_page)
        self.pushButton.clicked.connect(self.google_plus)
        self.pushButton_4.clicked.connect(self.github)


    def facebook_page(self):
        url = 'https://www.facebook.com/profile.php?id=100007110228791'
        webbrowser.open(url)

    def google_plus(self):
        url = 'https://plus.google.com/113481116939970297334?hl=ar'
        webbrowser.open(url)


    def github(self):
        url = 'https://github.com/ali-falah'
        webbrowser.open(url)

#######################################################
app=QApplication(sys.argv)
win=main()
win.statusBar().showMessage('Developed By Ali Falah')
win.show()
app.exec_()
